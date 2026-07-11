from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.views.decorators.csrf import csrf_exempt
from django.conf import settings
from django.http import HttpResponse
from django.db.models import Count, Q, Avg, Sum, Max
from django.utils import timezone
from datetime import date, timedelta, datetime
import json, csv, os
import io
from openpyxl import load_workbook
from .models import UserProfile, DailyTopic, CheckInRecord, Message, Reply, Notification, QuizQuestion, QuizResult, Announcement
from django.contrib.auth.models import User


# ── Admin access decorator ─────────────────────────────────────────────

def get_user_from_token(request):
    """从请求中获取Token对应的用户"""
    auth_header = request.META.get('HTTP_AUTHORIZATION', '')
    if auth_header.startswith('Token '):
        token_key = auth_header.split(' ')[1]
        from rest_framework.authtoken.models import Token
        try:
            token = Token.objects.get(key=token_key)
            return token.user
        except Token.DoesNotExist:
            pass
    return None


def admin_required(view_func):
    def wrapper(request, *args, **kwargs):
        user = request.user
        if user.is_authenticated:
            if user.is_staff:
                return view_func(request, *args, **kwargs)
            try:
                if user.profile.role in ['admin', 'super_admin']:
                    return view_func(request, *args, **kwargs)
            except:
                pass
        else:
            token_user = get_user_from_token(request)
            if token_user:
                user = token_user
                try:
                    if user.profile.role in ['admin', 'super_admin']:
                        return view_func(request, *args, **kwargs)
                except:
                    pass
        return redirect(settings.LOGIN_URL)
    return wrapper


def _get_managed_school(request):
    """获取当前管理员负责的学校。super_admin 返回 None 表示不受限制"""
    user = request.user
    if user.is_staff or (hasattr(user, 'profile') and user.profile.role == 'super_admin'):
        return None  # 超级管理员 / Django staff → 不受限制
    if hasattr(user, 'profile') and user.profile.role == 'admin':
        return user.profile.managed_school or '__none__'  # __none__ 表示没有分配学校，查不到任何用户
    return '__none__'


def _filter_profiles_by_school(request):
    """根据管理员学校过滤 UserProfile 查询集，排除管理员账号"""
    school = _get_managed_school(request)
    if school is None:
        return UserProfile.objects.exclude(role__in=['admin', 'super_admin'])
    if school == '__none__':
        return UserProfile.objects.none()
    return UserProfile.objects.filter(school=school).exclude(role__in=['admin', 'super_admin'])


def super_admin_required(view_func):
    def wrapper(request, *args, **kwargs):
        user = request.user
        if user.is_authenticated:
            if user.is_staff:
                return view_func(request, *args, **kwargs)
            try:
                if user.profile.role == 'super_admin':
                    return view_func(request, *args, **kwargs)
            except:
                pass
        else:
            token_user = get_user_from_token(request)
            if token_user:
                user = token_user
                try:
                    if user.profile.role == 'super_admin':
                        return view_func(request, *args, **kwargs)
                except:
                    pass
        return redirect(settings.LOGIN_URL)
    return wrapper


def _is_super_admin(user):
    """判断用户是否为超级管理员"""
    if not user.is_authenticated:
        return False
    if user.is_staff:
        return True
    return getattr(user.profile, 'role', '') == 'super_admin'


# ── Config helpers ─────────────────────────────────────────────────────

CONFIG_FILE = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'admin_config.json')

DEFAULT_CONFIG = {
    'activity_name': '21天安全用网打卡活动',
    'slogan': '每天打卡，做网络安全小达人！',
    'start_date': '',
    'end_date': '',
    'allow_makeup': False,
    'makeup_limit': 3,
}


def _load_config():
    try:
        with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
            config = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        config = {}
    return {**DEFAULT_CONFIG, **config}


def _save_config(data):
    config = {**DEFAULT_CONFIG, **data}
    with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
        json.dump(config, f, ensure_ascii=False, indent=2)


# ── Helper: calculate streak days (consecutive from day 1) ─────────────

def _calc_streak(user):
    """Count consecutive check-in days starting from day 1."""
    days = (
        CheckInRecord.objects.filter(user=user)
        .order_by('day')
        .values_list('day', flat=True)
    )
    if not days:
        return 0
    day_set = set(days)
    streak = 0
    for d in range(1, 22):
        if d in day_set:
            streak += 1
        else:
            break
    return streak


# ═══════════════════════════════════════════════════════════════════════
# 1. Admin Dashboard
# ═══════════════════════════════════════════════════════════════════════

@admin_required
def admin_index_view(request):
    return redirect('admin_dashboard')


@csrf_exempt
def admin_login_view(request):
    """管理员登录页面"""
    error = ''
    if request.method == 'POST':
        phone = request.POST.get('phone', '').strip()
        password = request.POST.get('password', '').strip()
        
        try:
            profile = UserProfile.objects.get(phone=phone)
            user = profile.user
            if user.check_password(password):
                if profile.role in ['admin', 'super_admin'] or user.is_staff:
                    # 设置 user.backend，否则 login() 无法持久化认证状态
                    user.backend = 'django.contrib.auth.backends.ModelBackend'
                    from django.contrib.auth import login as django_login
                    django_login(request, user)
                    return redirect('admin_dashboard')
                else:
                    error = '该账号不是管理员'
            else:
                error = '密码错误'
        except UserProfile.DoesNotExist:
            error = '手机号未注册'
    
    return render(request, 'admin/manage_login.html', {'error': error})


@admin_required
def admin_dashboard_view(request):
    # ── 获取有权限的学校用户 ──
    managed_school = _get_managed_school(request)
    if managed_school is None:
        profiles = UserProfile.objects.exclude(role__in=['admin', 'super_admin'])
        checkins = CheckInRecord.objects.exclude(user__profile__role__in=['admin', 'super_admin'])
    elif managed_school == '__none__':
        profiles = UserProfile.objects.none()
        checkins = CheckInRecord.objects.none()
    else:
        profiles = UserProfile.objects.filter(school=managed_school).exclude(role__in=['admin', 'super_admin'])
        checkins = CheckInRecord.objects.filter(user__profile__school=managed_school).exclude(user__profile__role__in=['admin', 'super_admin'])

    # ── Basic counts ──
    total_participants = profiles.count()
    total_checkins = checkins.count()
    overall_completion_rate = (
        round(total_checkins / (total_participants * 21) * 100, 1)
        if total_participants > 0 else 0
    )

    # ── Daily check-in counts (bar chart data) ──
    daily_stats = []
    for d in range(1, 22):
        cnt = checkins.filter(day=d).count()
        daily_stats.append({'day': d, 'count': cnt})

    # ── Duration distribution ──
    duration_labels = ['0-1小时', '1-3小时', '3-5小时', '5小时以上']
    duration_data = {}
    for label in duration_labels:
        duration_data[label] = checkins.filter(online_duration=label).count()
    duration_distribution = [
        {'label': label, 'count': duration_data[label], 'percent': round(duration_data[label] / max(total_checkins, 1) * 100, 1)}
        for label in duration_labels
    ]

    # ── Activity distribution ──
    activity_counter = {}
    for r in checkins:
        if r.online_activities:
            for act in r.online_activities.replace('，', ',').split(','):
                act = act.strip()
                if act:
                    activity_counter[act] = activity_counter.get(act, 0) + 1
    activity_distribution = sorted(
        [{'name': k, 'count': v} for k, v in activity_counter.items()],
        key=lambda x: x['count'], reverse=True
    )

    # ── Completed all 21 days ──
    completed_all = profiles.filter(
        user__checkins__day=21
    ).distinct().count()

    # ── Active users today ──
    today = date.today()
    active_today = checkins.filter(date=today).count()

    # ── Average screen time ──
    hour_map = {'0-1小时': 0.5, '1-3小时': 2, '3-5小时': 4, '5小时以上': 6}
    total_hours = 0
    for r in checkins:
        total_hours += hour_map.get(r.online_duration.strip() if r.online_duration else '', 0)
    avg_screen_time = round(total_hours / max(total_checkins, 1), 1)

    # ── Top activity ──
    top_activity = activity_distribution[0]['name'] if activity_distribution else '暂无数据'

    activity_distribution_with_index = []
    for idx, act in enumerate(activity_distribution):
        activity_distribution_with_index.append({**act, 'index': idx})

    context = {
        'current_page': 'dashboard',
        'total_participants': total_participants,
        'total_checkins': total_checkins,
        'overall_completion_rate': overall_completion_rate,
        'daily_stats': daily_stats,
        'duration_distribution': duration_distribution,
        'activity_distribution': activity_distribution_with_index,
        'completed_all': completed_all,
        'active_today': active_today,
        'avg_screen_time': avg_screen_time,
        'top_activity': top_activity,
        'today': datetime.now().strftime('%Y年%m月%d日'),
    }
    return render(request, 'admin/dashboard.html', context)


# ═══════════════════════════════════════════════════════════════════════
# 1.5 Export Check-in Data (Excel)
# ═══════════════════════════════════════════════════════════════════════

@admin_required
def admin_export_checkins(request):
    """导出管辖范围内的打卡数据为 Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    managed_school = _get_managed_school(request)
    if managed_school == '__none__':
        checkins = CheckInRecord.objects.none()
    elif managed_school is None:
        checkins = CheckInRecord.objects.exclude(
            user__profile__role__in=['admin', 'super_admin']
        ).select_related('user__profile').order_by('user__profile__school', 'user__profile__real_name', 'day')
    else:
        checkins = CheckInRecord.objects.filter(
            user__profile__school=managed_school
        ).exclude(
            user__profile__role__in=['admin', 'super_admin']
        ).select_related('user__profile').order_by('user__profile__real_name', 'day')

    wb = Workbook()
    ws = wb.active
    ws.title = '打卡数据'

    # ── Styles ──
    header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='6366F1', end_color='6366F1', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB'),
    )

    # ── Headers ──
    headers = [
        '序号', '手机号', '姓名', '学校', '年级',
        '打卡天数', '打卡日期', '学习时长(分)', '娱乐时长(分)',
        '上网活动', '心情', '自评(分)', '打卡时间',
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # ── Data rows ──
    for i, record in enumerate(checkins, start=1):
        profile = record.user.profile
        row_data = [
            i,
            profile.phone,
            profile.real_name or profile.nickname,
            profile.school or '-',
            profile.grade or '-',
            record.day,
            record.date.strftime('%Y-%m-%d') if record.date else '-',
            record.study_duration or 0,
            record.entertainment_duration or 0,
            record.online_activities or '-',
            record.mood or '-',
            record.self_rating or '-',
            record.created_at.strftime('%Y-%m-%d %H:%M') if record.created_at else '-',
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i + 1, column=col, value=val)
            cell.alignment = cell_align
            cell.border = thin_border

    # ── Column widths ──
    col_widths = [6, 16, 10, 12, 8, 8, 14, 14, 14, 28, 10, 10, 20]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

    # ── Freeze header ──
    ws.freeze_panes = 'A2'

    # ── Response ──
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    school_label = managed_school if managed_school else '全部'
    filename = f'打卡数据_{school_label}_{date.today().strftime("%Y%m%d")}.xlsx'
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response['Content-Length'] = len(response.content)
    return response


# ═══════════════════════════════════════════════════════════════════════
# 2. User Management
# ═══════════════════════════════════════════════════════════════════════

@admin_required
def admin_users_view(request):
    # ── CSV Export ──
    if request.GET.get('export') == 'csv':
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = 'attachment; filename="users_export.csv"'
        writer = csv.writer(response)
        writer.writerow(['手机号', '昵称', '学校', '年级', '总打卡天数', '完成进度(%)', '最后打卡日期'])

        profiles = _filter_profiles_by_school(request)
        for p in profiles:
            total = CheckInRecord.objects.filter(user=p.user).count()
            progress = round(total / 21 * 100, 1)
            last = CheckInRecord.objects.filter(user=p.user).order_by('-day').first()
            last_date = last.date.strftime('%Y-%m-%d') if last else '-'
            writer.writerow([p.phone, p.nickname, p.school, p.grade, total, progress, last_date])
        return response

    users_data = _build_users_data(request)

    total_users = len(users_data)
    completed_21 = sum(1 for u in users_data if u['total_checkins'] >= 21)
    active_streak = sum(1 for u in users_data if u['streak_days'] >= 7)
    today = date.today().strftime('%Y-%m-%d')
    checked_today = sum(1 for u in users_data if u['last_checkin_date'] == today)

    context = {
        'current_page': 'users',
        'users_data': users_data,
        'total_users': total_users,
        'completed_21': completed_21,
        'active_streak': active_streak,
        'checked_today': checked_today,
        'is_super_admin': _is_super_admin(request.user),
    }
    return render(request, 'admin/users.html', context)


# ═══════════════════════════════════════════════════════════════════════
# 2.5 Excel Import Users
# ═══════════════════════════════════════════════════════════════════════

@admin_required
def admin_import_users(request):
    """Excel 批量导入用户"""
    if request.method != 'POST':
        return redirect('admin_users')

    excel_file = request.FILES.get('file')
    if not excel_file:
        context = {
            'current_page': 'users',
            'users_data': _build_users_data(request),
            'import_error': '请选择一个 Excel 文件',
        }
        return render(request, 'admin/users.html', context)

    # 检查文件扩展名
    filename = excel_file.name.lower()
    if not (filename.endswith('.xlsx') or filename.endswith('.xls')):
        context = {
            'current_page': 'users',
            'users_data': _build_users_data(request),
            'import_error': '仅支持 .xlsx / .xls 格式的 Excel 文件',
        }
        return render(request, 'admin/users.html', context)

    try:
        wb = load_workbook(excel_file, read_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(min_row=2, values_only=True))  # 跳过表头
        imported = 0
        skipped = 0
        errors = []

        for i, row in enumerate(rows, start=2):
            # 读取列：学校, 真实姓名, 电话号
            school = str(row[0]).strip() if row[0] else ''
            real_name = str(row[1]).strip() if row[1] else ''
            phone = str(row[2]).strip() if row[2] else ''

            if not phone or not real_name or not school:
                errors.append(f'第{i}行：学校、姓名、电话号不能为空')
                continue

            # 跳过已存在的手机号
            if UserProfile.objects.filter(phone=phone).exists():
                skipped += 1
                continue

            try:
                user = User.objects.create_user(
                    username=phone,
                    password='123456',
                )
                UserProfile.objects.create(
                    user=user,
                    phone=phone,
                    nickname=real_name,
                    real_name=real_name,
                    school=school,
                    role='user',
                )
                imported += 1
            except Exception as e:
                errors.append(f'第{i}行({real_name})：创建失败 - {str(e)}')

        wb.close()

        msg = f'导入完成：成功 {imported} 人，跳过 {skipped} 人（已存在）'
        if errors:
            msg += f'，{len(errors)} 条错误'

        context = {
            'current_page': 'users',
            'users_data': _build_users_data(request),
            'import_success': msg,
            'import_errors': errors[:20],  # 最多显示 20 条错误
        }
        return render(request, 'admin/users.html', context)

    except Exception as e:
        context = {
            'current_page': 'users',
            'users_data': _build_users_data(request),
            'import_error': f'解析 Excel 文件失败：{str(e)}',
        }
        return render(request, 'admin/users.html', context)


def _build_users_data(request):
    """构建用户数据列表，供 users 页面和导入页面复用"""
    profiles = _filter_profiles_by_school(request)
    users_data = []
    for p in profiles:
        last_checkin = CheckInRecord.objects.filter(user=p.user).order_by('-created_at').first()
        total_checkins = CheckInRecord.objects.filter(user=p.user).count()
        streak_days = 0
        if total_checkins > 0:
            checkin_days = sorted(
                CheckInRecord.objects.filter(user=p.user).values_list('day', flat=True).distinct()
            )
            streak_days = 1
            for j in range(len(checkin_days) - 1, 0, -1):
                if checkin_days[j] == checkin_days[j - 1] + 1:
                    streak_days += 1
                else:
                    break
        users_data.append({
            'profile': p,
            'total_checkins': total_checkins,
            'streak_days': streak_days,
            'last_checkin_date': last_checkin.created_at.strftime('%Y-%m-%d') if last_checkin else '-',
        })
    return users_data


# ═══════════════════════════════════════════════════════════════════════
# 3. Daily Topic Management
# ═══════════════════════════════════════════════════════════════════════

@admin_required
def admin_topics_view(request):
    if request.method == 'POST':
        day = int(request.POST.get('day', 0))
        if 1 <= day <= 21:
            topic = DailyTopic.objects.filter(day=day).first()
            if topic:
                topic.title = request.POST.get('title', topic.title)
                topic.content = request.POST.get('content', topic.content)
                topic.question = request.POST.get('question', topic.question)
                topic.icon = request.POST.get('icon', topic.icon)
                topic.save()
        return redirect('admin_topics')

    # ── Drill-down by day ──
    day_filter = request.GET.get('day')
    day_records = None
    if day_filter:
        try:
            day_int = int(day_filter)
            allowed_users = _filter_profiles_by_school(request).values_list('user_id', flat=True)
            day_records = CheckInRecord.objects.filter(day=day_int, user_id__in=allowed_users).select_related('user__profile').order_by('-created_at')
        except ValueError:
            pass

    topics = DailyTopic.objects.all().order_by('day')
    topics_data = []
    for t in topics:
        answer_count = CheckInRecord.objects.filter(day=t.day).exclude(user__profile__role__in=['admin', 'super_admin']).count()
        topics_data.append({
            'topic': t,
            'answer_count': answer_count,
        })

    context = {
        'current_page': 'topics',
        'topics_data': topics_data,
        'day_records': day_records,
        'filter_day': day_filter,
    }
    return render(request, 'admin/topics.html', context)


# ═══════════════════════════════════════════════════════════════════════
# 4. Message Moderation
# ═══════════════════════════════════════════════════════════════════════

@admin_required
def admin_messages_view(request):
    if request.method == 'POST':
        action = request.POST.get('action', '')
        msg_id = request.POST.get('id', '')

        if msg_id:
            try:
                msg = Message.objects.get(id=int(msg_id))
            except (Message.DoesNotExist, ValueError):
                msg = None

            if msg:
                if action == 'delete':
                    msg.delete()
                elif action == 'hide':
                    if not msg.content.startswith('[已隐藏]'):
                        msg.content = '[已隐藏] ' + msg.content
                        msg.save()

        return redirect('admin_messages')

    # ── GET: list all messages ──
    allowed_users = _filter_profiles_by_school(request).values_list('user_id', flat=True)
    messages = Message.objects.filter(user_id__in=allowed_users).select_related('user__profile').prefetch_related('likes', 'replies').order_by('-created_at')

    msg_list = []
    for m in messages:
        try:
            nickname = m.user.profile.nickname
            phone = m.user.profile.phone
        except UserProfile.DoesNotExist:
            nickname = m.user.username
            phone = '-'
        msg_list.append({
            'message': m,
            'nickname': nickname,
            'phone': phone,
        })

    context = {
        'current_page': 'messages',
        'messages': msg_list,
    }
    return render(request, 'admin/messages.html', context)


# ═══════════════════════════════════════════════════════════════════════
# 5. Rule Configuration
# ═══════════════════════════════════════════════════════════════════════

@admin_required
def admin_config_view(request):
    if request.method == 'POST':
        config = {
            'activity_name': request.POST.get('activity_name', DEFAULT_CONFIG['activity_name']),
            'slogan': request.POST.get('slogan', DEFAULT_CONFIG['slogan']),
            'start_date': request.POST.get('start_date', ''),
            'end_date': request.POST.get('end_date', ''),
            'allow_makeup': request.POST.get('allow_makeup') == 'true',
            'makeup_limit': int(request.POST.get('makeup_limit', DEFAULT_CONFIG['makeup_limit'])),
        }
        _save_config(config)
        return redirect('admin_config')

    context = {
        'current_page': 'config',
        'config': _load_config(),
    }
    return render(request, 'admin/config.html', context)


# ═══════════════════════════════════════════════════════════════════════
# 6. Announcement Publisher
# ═══════════════════════════════════════════════════════════════════════

@admin_required
def admin_announce_view(request):
    """公告管理 - 发布/编辑/删除公告，可见范围与管理员权限一致"""
    from day21.views import create_notification

    # ── DELETE ──
    if request.GET.get('delete'):
        try:
            a = Announcement.objects.get(id=request.GET['delete'])
            a.delete()
        except Announcement.DoesNotExist:
            pass
        return redirect('admin_announce')

    # ── EDIT (load existing) ──
    edit_announcement = None
    if request.GET.get('edit'):
        try:
            edit_announcement = Announcement.objects.get(id=request.GET['edit'])
        except (Announcement.DoesNotExist, ValueError):
            pass

    # ── POST: Create or Update ──
    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        content = request.POST.get('content', '').strip()
        n_type = request.POST.get('n_type', 'announcement')
        edit_id = request.POST.get('edit_id', '')

        if title and content:
            allowed_users_qs = _filter_profiles_by_school(request)
            allowed_users = list(allowed_users_qs.values_list('user_id', flat=True))
            users = User.objects.filter(id__in=allowed_users)

            # 确定目标学校
            managed_school = _get_managed_school(request)
            target_school = managed_school if managed_school and managed_school != '__none__' else ''

            if edit_id:
                # 更新已有公告
                try:
                    a = Announcement.objects.get(id=int(edit_id))
                    a.title = title
                    a.content = content
                    a.n_type = n_type
                    a.target_school = target_school
                    a.recipient_count = len(allowed_users)
                    a.save()
                except Announcement.DoesNotExist:
                    a = Announcement.objects.create(
                        title=title, content=content, n_type=n_type,
                        target_school=target_school,
                        publisher=request.user,
                        recipient_count=len(allowed_users),
                    )
            else:
                a = Announcement.objects.create(
                    title=title, content=content, n_type=n_type,
                    target_school=target_school,
                    publisher=request.user,
                    recipient_count=len(allowed_users),
                )

            for user in users:
                create_notification(user, title, content, n_type)

        return redirect('admin_announce')

    # ── GET: List announcements ──
    managed_school = _get_managed_school(request)
    if managed_school is None:
        announcements = Announcement.objects.all().order_by('-created_at')
        scope_desc = '所有用户'
    elif managed_school == '__none__':
        announcements = Announcement.objects.none()
        scope_desc = '无权限'
    else:
        announcements = Announcement.objects.filter(target_school=managed_school).order_by('-created_at')
        scope_desc = f'学校: {managed_school}'

    context = {
        'current_page': 'announce',
        'announcements': announcements,
        'edit_announcement': edit_announcement,
        'scope_desc': scope_desc,
    }
    return render(request, 'admin/announce.html', context)


# ═══════════════════════════════════════════════════════════════════════
# 7. Quiz Management
# ═══════════════════════════════════════════════════════════════════════

@admin_required
def admin_quiz_view(request):
    """题库管理"""
    import random

    # ── BULK DELETE ──
    if request.GET.get('bulk_delete'):
        ids = request.GET.getlist('ids')
        if ids:
            QuizQuestion.objects.filter(id__in=ids).delete()
        return redirect('admin_quiz')

    # ── POST: Create / Update ──
    if request.method == 'POST':
        qid = request.POST.get('id')
        q_type = request.POST.get('q_type', 'choice')

        if qid:
            q = QuizQuestion.objects.get(id=qid)
            q.q_type = q_type
            q.day = int(request.POST.get('day', q.day))
            q.question = request.POST.get('question', q.question)
            q.answer = request.POST.get('answer', q.answer)
            q.explanation = request.POST.get('explanation', q.explanation)
            if q_type == 'true_false':
                q.option_a = '正确'
                q.option_b = '错误'
                q.option_c = ''
                q.option_d = ''
            else:
                q.option_a = request.POST.get('option_a', '')
                q.option_b = request.POST.get('option_b', '')
                q.option_c = request.POST.get('option_c', '')
                q.option_d = request.POST.get('option_d', '')
            q.save()
        else:
            # Add new
            kw = {
                'q_type': q_type,
                'day': int(request.POST.get('day', 1)),
                'question': request.POST.get('question', ''),
                'answer': request.POST.get('answer', 'A'),
                'explanation': request.POST.get('explanation', ''),
            }
            if q_type == 'true_false':
                kw['option_a'] = '正确'
                kw['option_b'] = '错误'
                kw['option_c'] = ''
                kw['option_d'] = ''
            else:
                kw['option_a'] = request.POST.get('option_a', '')
                kw['option_b'] = request.POST.get('option_b', '')
                kw['option_c'] = request.POST.get('option_c', '')
                kw['option_d'] = request.POST.get('option_d', '')
            QuizQuestion.objects.create(**kw)
        return redirect('admin_quiz')

    # ── DELETE ──
    if request.GET.get('delete'):
        QuizQuestion.objects.filter(id=request.GET['delete']).delete()
        return redirect('admin_quiz')

    # ── EXPORT ──
    if request.GET.get('export') == '1':
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '题库数据'

        # Styled header
        header_font = openpyxl.styles.Font(bold=True, color='FFFFFF', size=11)
        header_fill = openpyxl.styles.PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
        header_align = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        thin_border = openpyxl.styles.Border(
            left=openpyxl.styles.Side(style='thin', color='D1D5DB'),
            right=openpyxl.styles.Side(style='thin', color='D1D5DB'),
            top=openpyxl.styles.Side(style='thin', color='D1D5DB'),
            bottom=openpyxl.styles.Side(style='thin', color='D1D5DB'),
        )

        headers = ['ID', '题型', '天数', '题目', '选项A', '选项B', '选项C', '选项D', '答案', '解析']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 8
        ws.column_dimensions['D'].width = 55
        ws.column_dimensions['E'].width = 28
        ws.column_dimensions['F'].width = 28
        ws.column_dimensions['G'].width = 28
        ws.column_dimensions['H'].width = 28
        ws.column_dimensions['I'].width = 8
        ws.column_dimensions['J'].width = 40

        qs = QuizQuestion.objects.all().order_by('day', 'id')
        for row, q in enumerate(qs, 2):
            data = [
                q.id,
                '判断题' if q.q_type == 'true_false' else '选择题',
                q.day,
                q.question,
                q.option_a or '',
                q.option_b or '',
                q.option_c or '',
                q.option_d or '',
                q.answer,
                q.explanation or '',
            ]
            for col, val in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = thin_border
                cell.alignment = openpyxl.styles.Alignment(vertical='center', wrap_text=True)

        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=quiz_export.xlsx'
        wb.save(response)
        return response

    # ── IMPORT ──
    if request.method == 'POST' and request.FILES.get('import_file'):
        import openpyxl
        from io import BytesIO

        file = request.FILES['import_file']
        ext = file.name.rsplit('.', 1)[-1].lower()

        rows = []
        errors = []

        if ext in ('xlsx', 'xls'):
            wb = openpyxl.load_workbook(BytesIO(file.read()))
            ws = wb.active
            raw = list(ws.iter_rows(values_only=True))
            if not raw:
                errors.append('文件为空')
            else:
                header = raw[0]
                header_map = {}
                for i, h in enumerate(header):
                    if h:
                        hh = str(h).strip()
                        if '题型' in hh:
                            header_map['q_type'] = i
                        elif hh in ('天数', 'day', 'Day'):
                            header_map['day'] = i
                        elif '题目' in hh or '问题' in hh or '题干' in hh:
                            header_map['question'] = i
                        elif '选项A' in hh or 'A' == hh:
                            header_map['option_a'] = i
                        elif '选项B' in hh or 'B' == hh:
                            header_map['option_b'] = i
                        elif '选项C' in hh or 'C' == hh:
                            header_map['option_c'] = i
                        elif '选项D' in hh or 'D' == hh:
                            header_map['option_d'] = i
                        elif '答案' in hh or '正确' in hh:
                            header_map['answer'] = i
                        elif '解析' in hh or '解释' in hh:
                            header_map['explanation'] = i

                for row_idx, row in enumerate(raw[1:], 2):
                    def col(key):
                        idx = header_map.get(key)
                        return str(row[idx]).strip() if idx is not None and row[idx] is not None else ''

                    q_type = col('q_type')
                    question = col('question')
                    day_val = col('day')
                    answer = col('answer')

                    if not question:
                        continue
                    if not day_val or not day_val.isdigit():
                        errors.append(f'第{row_idx}行: 天数无效')
                        continue

                    q_type = 'true_false' if '判断' in q_type else 'choice'
                    answer = answer.upper()
                    if answer not in ('A', 'B', 'C', 'D'):
                        errors.append(f'第{row_idx}行: 答案格式无效({answer})')
                        continue

                    if q_type == 'true_false':
                        rows.append({
                            'q_type': 'true_false',
                            'day': int(day_val),
                            'question': question,
                            'option_a': '正确',
                            'option_b': '错误',
                            'option_c': '',
                            'option_d': '',
                            'answer': answer,
                            'explanation': col('explanation'),
                        })
                    else:
                        rows.append({
                            'q_type': 'choice',
                            'day': int(day_val),
                            'question': question,
                            'option_a': col('option_a'),
                            'option_b': col('option_b'),
                            'option_c': col('option_c'),
                            'option_d': col('option_d'),
                            'answer': answer,
                            'explanation': col('explanation'),
                        })

        elif ext == 'json':
            import json
            data = json.loads(file.read().decode('utf-8'))
            if isinstance(data, dict):
                data = [data]
            for item in data:
                q_type = item.get('q_type', 'choice')
                if '判断' in str(q_type):
                    q_type = 'true_false'
                elif '选择' in str(q_type):
                    q_type = 'choice'

                answer = str(item.get('answer', 'A')).upper()
                if answer not in ('A', 'B', 'C', 'D'):
                    errors.append(f'题目"{item.get("question","")[:20]}..."答案无效')
                    continue

                day_val = item.get('day', 1)
                try:
                    day_val = int(day_val)
                except (ValueError, TypeError):
                    day_val = 1

                if q_type == 'true_false':
                    rows.append({
                        'q_type': 'true_false',
                        'day': day_val,
                        'question': str(item.get('question', '')),
                        'option_a': '正确',
                        'option_b': '错误',
                        'option_c': '',
                        'option_d': '',
                        'answer': answer,
                        'explanation': str(item.get('explanation', '')),
                    })
                else:
                    rows.append({
                        'q_type': 'choice',
                        'day': day_val,
                        'question': str(item.get('question', '')),
                        'option_a': str(item.get('option_a', '')),
                        'option_b': str(item.get('option_b', '')),
                        'option_c': str(item.get('option_c', '')),
                        'option_d': str(item.get('option_d', '')),
                        'answer': answer,
                        'explanation': str(item.get('explanation', '')),
                    })
        else:
            errors.append(f'不支持的文件格式: .{ext}')

        created = 0
        skipped = 0
        for r in rows:
            if QuizQuestion.objects.filter(question=r['question'], day=r['day']).exists():
                skipped += 1
                continue
            QuizQuestion.objects.create(**r)
            created += 1

        # Store results in session for display
        request.session['import_result'] = {
            'created': created,
            'skipped': skipped,
            'errors': errors,
        }
        return redirect('admin_quiz')

    # ── Get import result from session ──
    import_result = request.session.pop('import_result', None)

    # ── Filter & Search ──
    questions = QuizQuestion.objects.all()
    day_filter = request.GET.get('day', '')
    type_filter = request.GET.get('q_type', '')
    search = request.GET.get('search', '').strip()

    if day_filter:
        questions = questions.filter(day=int(day_filter))
    if type_filter:
        questions = questions.filter(q_type=type_filter)
    if search:
        questions = questions.filter(question__icontains=search)

    questions = questions.order_by('day', 'id')

    edit_q = None
    if request.GET.get('edit'):
        eid = request.GET['edit']
        if eid == '0':
            edit_q = None
        else:
            try:
                edit_q = QuizQuestion.objects.get(id=eid)
            except QuizQuestion.DoesNotExist:
                pass

    days = sorted(set(QuizQuestion.objects.values_list('day', flat=True)))
    day_info = []
    for d in days:
        day_info.append({
            'day': d,
            'count': QuizQuestion.objects.filter(day=d).count(),
            'choice_count': QuizQuestion.objects.filter(day=d, q_type='choice').count(),
            'tf_count': QuizQuestion.objects.filter(day=d, q_type='true_false').count(),
        })

    context = {
        'current_page': 'quiz',
        'import_result': import_result,
        'questions': questions,
        'edit_q': edit_q,
        'day_filter': int(day_filter) if day_filter else '',
        'type_filter': type_filter,
        'search': search,
        'days': days,
        'day_info': day_info,
    }
    return render(request, 'admin/quiz.html', context)


# ═══════════════════════════════════════════════════════════════════════
# 8. Real-name Authentication Review
# ═══════════════════════════════════════════════════════════════════════

@admin_required
def admin_auth_view(request):
    """实名认证审核"""
    if request.method == 'POST':
        action = request.POST.get('action', '')
        profile_id = request.POST.get('profile_id', '')
        reason = request.POST.get('reason', '')

        if profile_id:
            try:
                profile = UserProfile.objects.get(id=int(profile_id))
            except (UserProfile.DoesNotExist, ValueError):
                profile = None

            if profile:
                if action == 'approve':
                    profile.auth_status = 'approved'
                    profile.auth_reason = reason
                    profile.auth_time = timezone.now()
                    profile.auth_operator = request.user
                    profile.save()

                    Notification.objects.create(
                        user=profile.user,
                        title='实名认证审核通过',
                        content='恭喜您！实名认证已通过审核',
                        type='auth',
                    )

                elif action == 'reject':
                    profile.auth_status = 'rejected'
                    profile.auth_reason = reason
                    profile.auth_time = timezone.now()
                    profile.auth_operator = request.user
                    profile.save()

                    Notification.objects.create(
                        user=profile.user,
                        title='实名认证审核未通过',
                        content=f'您的实名认证申请未通过，原因：{reason}',
                        type='auth',
                    )

        return redirect('admin_auth')

    status_filter = request.GET.get('status', '')
    profiles = _filter_profiles_by_school(request)
    if status_filter:
        profiles = profiles.filter(auth_status=status_filter)
    profiles = profiles.order_by('-auth_time', 'auth_status')

    auth_data = []
    for p in profiles:
        id_card_mask = p.id_card[:4] + '**********' + p.id_card[-4:] if p.id_card else ''
        total_checkins = CheckInRecord.objects.filter(user=p.user).count()
        auth_data.append({
            'profile': p,
            'id_card_mask': id_card_mask,
            'total_checkins': total_checkins,
        })

    pending_count = _filter_profiles_by_school(request).filter(auth_status='pending').count()
    approved_count = _filter_profiles_by_school(request).filter(auth_status='approved').count()
    rejected_count = _filter_profiles_by_school(request).filter(auth_status='rejected').count()
    unverified_count = _filter_profiles_by_school(request).filter(auth_status='unverified').count()

    context = {
        'current_page': 'auth',
        'auth_data': auth_data,
        'status_filter': status_filter,
        'pending_count': pending_count,
        'approved_count': approved_count,
        'rejected_count': rejected_count,
        'unverified_count': unverified_count,
    }
    return render(request, 'admin/auth.html', context)


@super_admin_required
def admin_admin_manage(request):
    """管理员管理（仅超级管理员）"""
    # 只有超级管理员才能访问
    user = request.user
    is_super = user.is_authenticated and (user.is_staff or getattr(user.profile, 'role', '') == 'super_admin')
    if not is_super:
        return redirect('admin_dashboard')
    
    admins = UserProfile.objects.filter(role__in=['admin', 'super_admin']).order_by('role', '-created_at')
    
    if request.method == 'POST':
        action = request.POST.get('action', '')
        profile_id = request.POST.get('profile_id', '')
        
        if action == 'add':
            phone = request.POST.get('phone', '').strip()
            nickname = request.POST.get('nickname', '').strip()
            password = request.POST.get('password', '').strip()
            role = request.POST.get('role', 'admin')
            
            if phone and nickname and password:
                if UserProfile.objects.filter(phone=phone).exists():
                    error = '该手机号已注册'
                else:
                    username = f'admin_{phone}'
                    new_user = User.objects.create_user(username=username, password=password)
                    UserProfile.objects.create(
                        user=new_user,
                        phone=phone,
                        nickname=nickname,
                        role=role,
                    )
                    return redirect('admin_admin_manage')
        
        elif action == 'edit':
            if profile_id:
                try:
                    profile = UserProfile.objects.get(id=int(profile_id))
                    # 不允许编辑超级管理员
                    if profile.role == 'super_admin':
                        return redirect('admin_admin_manage')
                    profile.nickname = request.POST.get('nickname', '').strip()
                    new_role = request.POST.get('role', 'admin')
                    # 只能设置为 admin 或 super_admin
                    if new_role in ['admin', 'super_admin']:
                        profile.role = new_role
                    profile.save()
                except UserProfile.DoesNotExist:
                    pass
            return redirect('admin_admin_manage')
        
        elif action == 'delete':
            if profile_id:
                try:
                    profile = UserProfile.objects.get(id=int(profile_id))
                    # 不能删除超级管理员
                    if profile.role != 'super_admin':
                        profile.user.delete()
                except UserProfile.DoesNotExist:
                    pass
            return redirect('admin_admin_manage')
    
    context = {
        'current_page': 'admin_manage',
        'admins': admins,
    }
    return render(request, 'admin/admin_manage.html', context)


# ═══════════════════════════════════════════════════════════════════════
# 11. User Account Management (Super Admin Only)
# ═══════════════════════════════════════════════════════════════════════

@super_admin_required
def admin_toggle_user(request, user_id):
    """启用/禁用用户账号"""
    try:
        profile = UserProfile.objects.get(id=user_id)
        user = profile.user
        user.is_active = not user.is_active
        user.save()
    except UserProfile.DoesNotExist:
        pass
    return redirect('admin_users')


@super_admin_required
def admin_reset_password(request, user_id):
    """重置用户密码为 123456"""
    try:
        profile = UserProfile.objects.get(id=user_id)
        user = profile.user
        user.set_password('123456')
        user.save()
    except UserProfile.DoesNotExist:
        pass
    return redirect('admin_users')
