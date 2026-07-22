"""Excel 导出公共工具 — 统一样式定义，消除 dashboard.py 和 quiz.py 的重复代码。"""
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from django.http import HttpResponse


# ── 默认样式 ──

DEFAULT_HEADER_FILL = '6366F1'          # 靛蓝色表头底色
DEFAULT_BORDER_COLOR = 'D1D5DB'         # 边框灰


def make_styles(header_color: str = DEFAULT_HEADER_FILL):
    """返回一组统一的 Excel 样式字典。"""
    return {
        'header_font': Font(name='微软雅黑', bold=True, size=11, color='FFFFFF'),
        'header_fill': PatternFill(start_color=header_color, end_color=header_color, fill_type='solid'),
        'header_align': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'cell_align': Alignment(horizontal='center', vertical='center'),
        'thin_border': Border(
            left=Side(style='thin', color=DEFAULT_BORDER_COLOR),
            right=Side(style='thin', color=DEFAULT_BORDER_COLOR),
            top=Side(style='thin', color=DEFAULT_BORDER_COLOR),
            bottom=Side(style='thin', color=DEFAULT_BORDER_COLOR),
        ),
    }


def write_header(ws, headers: list[str], styles: dict):
    """将表头写入第一行并应用样式。"""
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = styles['header_font']
        cell.fill = styles['header_fill']
        cell.alignment = styles['header_align']
        cell.border = styles['thin_border']


def write_row(ws, row_num: int, data: list, styles: dict):
    """将一行数据写入并应用单元格样式。"""
    for col, val in enumerate(data, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.alignment = styles['cell_align']
        cell.border = styles['thin_border']


def set_col_widths(ws, widths: list[int]):
    """按列索引设置列宽。"""
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w


def make_response(wb, filename: str) -> HttpResponse:
    """将 Workbook 保存为 HttpResponse 并返回。"""
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    resp['Content-Length'] = len(resp.content)
    return resp
