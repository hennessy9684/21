from django.shortcuts import render, redirect
from django.http import HttpResponse
from ..models import QuizQuestion, QuizResult
from .decorators import admin_required


# ═══════════════════════════════════════════════════════════════════════
# 7. Quiz Management
# ═══════════════════════════════════════════════════════════════════════

@admin_required
def admin_quiz_view(request):
    """题库管理"""
    import random

    # ── BULK DELETE ──
    if request.GET.get('bulk_delete'):
        ids = request.GET.getlist('ids')
        if ids:
            QuizQuestion.objects.filter(id__in=ids).delete()
        return redirect('admin_quiz')

    # ── POST: Create / Update ──
    if request.method == 'POST':
        qid = request.POST.get('id')
        q_type = request.POST.get('q_type', 'choice')

        if qid:
            q = QuizQuestion.objects.get(id=qid)
            q.q_type = q_type
            q.day = int(request.POST.get('day', q.day))
            q.question = request.POST.get('question', q.question)
            q.answer = request.POST.get('answer', q.answer)
            q.explanation = request.POST.get('explanation', q.explanation)
            if q_type == 'true_false':
                q.option_a = '正确'
                q.option_b = '错误'
                q.option_c = ''
                q.option_d = ''
            else:
                q.option_a = request.POST.get('option_a', '')
                q.option_b = request.POST.get('option_b', '')
                q.option_c = request.POST.get('option_c', '')
                q.option_d = request.POST.get('option_d', '')
            q.save()
        else:
            # Add new
            kw = {
                'q_type': q_type,
                'day': int(request.POST.get('day', 1)),
                'question': request.POST.get('question', ''),
                'answer': request.POST.get('answer', 'A'),
                'explanation': request.POST.get('explanation', ''),
            }
            if q_type == 'true_false':
                kw['option_a'] = '正确'
                kw['option_b'] = '错误'
                kw['option_c'] = ''
                kw['option_d'] = ''
            else:
                kw['option_a'] = request.POST.get('option_a', '')
                kw['option_b'] = request.POST.get('option_b', '')
                kw['option_c'] = request.POST.get('option_c', '')
                kw['option_d'] = request.POST.get('option_d', '')
            QuizQuestion.objects.create(**kw)
        return redirect('admin_quiz')

    # ── DELETE ──
    if request.GET.get('delete'):
        QuizQuestion.objects.filter(id=request.GET['delete']).delete()
        return redirect('admin_quiz')

    # ── EXPORT ──
    if request.GET.get('export') == '1':
        import openpyxl
        from .excel_utils import make_styles, write_header, set_col_widths, make_response

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '题库数据'
        styles = make_styles(header_color='4F46E5')

        headers = ['ID', '题型', '天数', '题目', '选项A', '选项B', '选项C', '选项D', '答案', '解析']
        write_header(ws, headers, styles)

        col_widths = [6, 10, 8, 55, 28, 28, 28, 28, 8, 40]
        set_col_widths(ws, col_widths)

        qs = QuizQuestion.objects.all().order_by('day', 'id')
        for row, q in enumerate(qs, 2):
            data = [
                q.id, '判断题' if q.q_type == 'true_false' else '选择题',
                q.day, q.question,
                q.option_a or '', q.option_b or '', q.option_c or '', q.option_d or '',
                q.answer, q.explanation or '',
            ]
            for col, val in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = styles['thin_border']
                cell.alignment = openpyxl.styles.Alignment(vertical='center', wrap_text=True)

        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions
        return make_response(wb, 'quiz_export.xlsx')

    # ── IMPORT ──
    if request.method == 'POST' and request.FILES.get('import_file'):
        import openpyxl
        from io import BytesIO

        file = request.FILES['import_file']
        ext = file.name.rsplit('.', 1)[-1].lower()

        rows = []
        errors = []

        if ext in ('xlsx', 'xls'):
            wb = openpyxl.load_workbook(BytesIO(file.read()))
            ws = wb.active
            raw = list(ws.iter_rows(values_only=True))
            if not raw:
                errors.append('文件为空')
            else:
                header = raw[0]
                header_map = {}
                for i, h in enumerate(header):
                    if h:
                        hh = str(h).strip()
                        if '题型' in hh:
                            header_map['q_type'] = i
                        elif hh in ('天数', 'day', 'Day'):
                            header_map['day'] = i
                        elif '题目' in hh or '问题' in hh or '题干' in hh:
                            header_map['question'] = i
                        elif '选项A' in hh or 'A' == hh:
                            header_map['option_a'] = i
                        elif '选项B' in hh or 'B' == hh:
                            header_map['option_b'] = i
                        elif '选项C' in hh or 'C' == hh:
                            header_map['option_c'] = i
                        elif '选项D' in hh or 'D' == hh:
                            header_map['option_d'] = i
                        elif '答案' in hh or '正确' in hh:
                            header_map['answer'] = i
                        elif '解析' in hh or '解释' in hh:
                            header_map['explanation'] = i

                for row_idx, row in enumerate(raw[1:], 2):
                    def col(key):
                        idx = header_map.get(key)
                        return str(row[idx]).strip() if idx is not None and row[idx] is not None else ''

                    q_type = col('q_type')
                    question = col('question')
                    day_val = col('day')
                    answer = col('answer')

                    if not question:
                        continue
                    if not day_val or not day_val.isdigit():
                        errors.append(f'第{row_idx}行: 天数无效')
                        continue

                    q_type = 'true_false' if '判断' in q_type else 'choice'
                    answer = answer.upper()
                    if answer not in ('A', 'B', 'C', 'D'):
                        errors.append(f'第{row_idx}行: 答案格式无效({answer})')
                        continue

                    if q_type == 'true_false':
                        rows.append({
                            'q_type': 'true_false',
                            'day': int(day_val),
                            'question': question,
                            'option_a': '正确',
                            'option_b': '错误',
                            'option_c': '',
                            'option_d': '',
                            'answer': answer,
                            'explanation': col('explanation'),
                        })
                    else:
                        rows.append({
                            'q_type': 'choice',
                            'day': int(day_val),
                            'question': question,
                            'option_a': col('option_a'),
                            'option_b': col('option_b'),
                            'option_c': col('option_c'),
                            'option_d': col('option_d'),
                            'answer': answer,
                            'explanation': col('explanation'),
                        })

        elif ext == 'json':
            import json
            data = json.loads(file.read().decode('utf-8'))
            if isinstance(data, dict):
                data = [data]
            for item in data:
                q_type = item.get('q_type', 'choice')
                if '判断' in str(q_type):
                    q_type = 'true_false'
                elif '选择' in str(q_type):
                    q_type = 'choice'

                answer = str(item.get('answer', 'A')).upper()
                if answer not in ('A', 'B', 'C', 'D'):
                    errors.append(f'题目"{item.get("question","")[:20]}..."答案无效')
                    continue

                day_val = item.get('day', 1)
                try:
                    day_val = int(day_val)
                except (ValueError, TypeError):
                    day_val = 1

                if q_type == 'true_false':
                    rows.append({
                        'q_type': 'true_false',
                        'day': day_val,
                        'question': str(item.get('question', '')),
                        'option_a': '正确',
                        'option_b': '错误',
                        'option_c': '',
                        'option_d': '',
                        'answer': answer,
                        'explanation': str(item.get('explanation', '')),
                    })
                else:
                    rows.append({
                        'q_type': 'choice',
                        'day': day_val,
                        'question': str(item.get('question', '')),
                        'option_a': str(item.get('option_a', '')),
                        'option_b': str(item.get('option_b', '')),
                        'option_c': str(item.get('option_c', '')),
                        'option_d': str(item.get('option_d', '')),
                        'answer': answer,
                        'explanation': str(item.get('explanation', '')),
                    })
        else:
            errors.append(f'不支持的文件格式: .{ext}')

        created = 0
        skipped = 0
        for r in rows:
            if QuizQuestion.objects.filter(question=r['question'], day=r['day']).exists():
                skipped += 1
                continue
            QuizQuestion.objects.create(**r)
            created += 1

        # Store results in session for display
        request.session['import_result'] = {
            'created': created,
            'skipped': skipped,
            'errors': errors,
        }
        return redirect('admin_quiz')

    # ── Get import result from session ──
    import_result = request.session.pop('import_result', None)

    # ── Filter & Search ──
    questions = QuizQuestion.objects.all()
    day_filter = request.GET.get('day', '')
    type_filter = request.GET.get('q_type', '')
    search = request.GET.get('search', '').strip()

    if day_filter:
        questions = questions.filter(day=int(day_filter))
    if type_filter:
        questions = questions.filter(q_type=type_filter)
    if search:
        questions = questions.filter(question__icontains=search)

    questions = questions.order_by('day', 'id')

    edit_q = None
    if request.GET.get('edit'):
        eid = request.GET['edit']
        if eid == '0':
            edit_q = None
        else:
            try:
                edit_q = QuizQuestion.objects.get(id=eid)
            except QuizQuestion.DoesNotExist:
                pass

    days = sorted(set(QuizQuestion.objects.values_list('day', flat=True)))
    day_info = []
    for d in days:
        day_info.append({
            'day': d,
            'count': QuizQuestion.objects.filter(day=d).count(),
            'choice_count': QuizQuestion.objects.filter(day=d, q_type='choice').count(),
            'tf_count': QuizQuestion.objects.filter(day=d, q_type='true_false').count(),
        })

    context = {
        'current_page': 'quiz',
        'import_result': import_result,
        'questions': questions,
        'edit_q': edit_q,
        'day_filter': int(day_filter) if day_filter else '',
        'type_filter': type_filter,
        'search': search,
        'days': days,
        'day_info': day_info,
    }
    return render(request, 'admin/quiz.html', context)
