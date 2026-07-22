from django.shortcuts import render, redirect
from django.views.decorators.csrf import csrf_exempt
from django.conf import settings
from django.http import HttpResponse
from datetime import date, datetime
import io
from ..models import UserProfile, CheckInRecord
from django.contrib.auth.models import User
from .decorators import admin_required, _get_managed_school


# ═══════════════════════════════════════════════════════════════════════
# 1. Admin Dashboard
# ═══════════════════════════════════════════════════════════════════════

@admin_required
def admin_index_view(request):
    return redirect('admin_dashboard')


@csrf_exempt
def admin_login_view(request):
    """管理员登录页面"""
    error = ''
    if request.method == 'POST':
        phone = request.POST.get('phone', '').strip()
        password = request.POST.get('password', '').strip()
        
        try:
            profile = UserProfile.objects.get(phone=phone)
            user = profile.user
            if user.check_password(password):
                if profile.role in ['admin', 'super_admin'] or user.is_staff:
                    # 设置 user.backend，否则 login() 无法持久化认证状态
                    user.backend = 'django.contrib.auth.backends.ModelBackend'
                    from django.contrib.auth import login as django_login
                    django_login(request, user)
                    return redirect('admin_dashboard')
                else:
                    error = '该账号不是管理员'
            else:
                error = '密码错误'
        except UserProfile.DoesNotExist:
            error = '手机号未注册'
    
    return render(request, 'admin/manage_login.html', {'error': error})


@admin_required
def admin_dashboard_view(request):
    # ── 获取有权限的学校用户 ──
    managed_school = _get_managed_school(request)
    if managed_school is None:
        profiles = UserProfile.objects.exclude(role__in=['admin', 'super_admin'])
        checkins = CheckInRecord.objects.exclude(user__profile__role__in=['admin', 'super_admin'])
    elif managed_school == '__none__':
        profiles = UserProfile.objects.none()
        checkins = CheckInRecord.objects.none()
    else:
        profiles = UserProfile.objects.filter(school=managed_school).exclude(role__in=['admin', 'super_admin'])
        checkins = CheckInRecord.objects.filter(user__profile__school=managed_school).exclude(user__profile__role__in=['admin', 'super_admin'])

    # ── Basic counts ──
    total_participants = profiles.count()
    total_checkins = checkins.count()
    overall_completion_rate = (
        round(total_checkins / (total_participants * 21) * 100, 1)
        if total_participants > 0 else 0
    )

    # ── Daily check-in counts (bar chart data) ──
    daily_stats = []
    for d in range(1, 22):
        cnt = checkins.filter(day=d).count()
        daily_stats.append({'day': d, 'count': cnt})

    # ── Duration distribution ──
    duration_labels = ['0-1小时', '1-3小时', '3-5小时', '5小时以上']
    duration_data = {}
    for label in duration_labels:
        duration_data[label] = checkins.filter(online_duration=label).count()
    duration_distribution = [
        {'label': label, 'count': duration_data[label], 'percent': round(duration_data[label] / max(total_checkins, 1) * 100, 1)}
        for label in duration_labels
    ]

    # ── Activity distribution ──
    activity_counter = {}
    for r in checkins:
        if r.online_activities:
            for act in r.online_activities.replace('，', ',').split(','):
                act = act.strip()
                if act:
                    activity_counter[act] = activity_counter.get(act, 0) + 1
    activity_distribution = sorted(
        [{'name': k, 'count': v} for k, v in activity_counter.items()],
        key=lambda x: x['count'], reverse=True
    )

    # ── Completed all 21 days ──
    completed_all = profiles.filter(
        user__checkins__day=21
    ).distinct().count()

    # ── Active users today ──
    today = date.today()
    active_today = checkins.filter(date=today).count()

    # ── Average screen time ──
    hour_map = {'0-1小时': 0.5, '1-3小时': 2, '3-5小时': 4, '5小时以上': 6}
    total_hours = 0
    for r in checkins:
        total_hours += hour_map.get(r.online_duration.strip() if r.online_duration else '', 0)
    avg_screen_time = round(total_hours / max(total_checkins, 1), 1)

    # ── Top activity ──
    top_activity = activity_distribution[0]['name'] if activity_distribution else '暂无数据'

    activity_distribution_with_index = []
    for idx, act in enumerate(activity_distribution):
        activity_distribution_with_index.append({**act, 'index': idx})

    context = {
        'current_page': 'dashboard',
        'total_participants': total_participants,
        'total_checkins': total_checkins,
        'overall_completion_rate': overall_completion_rate,
        'daily_stats': daily_stats,
        'duration_distribution': duration_distribution,
        'activity_distribution': activity_distribution_with_index,
        'completed_all': completed_all,
        'active_today': active_today,
        'avg_screen_time': avg_screen_time,
        'top_activity': top_activity,
        'today': datetime.now().strftime('%Y年%m月%d日'),
    }
    return render(request, 'admin/dashboard.html', context)


# ═══════════════════════════════════════════════════════════════════════
# 1.5 Export Check-in Data (Excel)
# ═══════════════════════════════════════════════════════════════════════

@admin_required
def admin_export_checkins(request):
    """导出管辖范围内的打卡数据为 Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    managed_school = _get_managed_school(request)
    if managed_school == '__none__':
        checkins = CheckInRecord.objects.none()
    elif managed_school is None:
        checkins = CheckInRecord.objects.exclude(
            user__profile__role__in=['admin', 'super_admin']
        ).select_related('user__profile').order_by('user__profile__school', 'user__profile__real_name', 'day')
    else:
        checkins = CheckInRecord.objects.filter(
            user__profile__school=managed_school
        ).exclude(
            user__profile__role__in=['admin', 'super_admin']
        ).select_related('user__profile').order_by('user__profile__real_name', 'day')

    wb = Workbook()
    ws = wb.active
    ws.title = '打卡数据'

    # ── Styles ──
    header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='6366F1', end_color='6366F1', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB'),
    )

    # ── Headers ──
    headers = [
        '序号', '手机号', '姓名', '学校', '年级',
        '打卡天数', '打卡日期', '学习时长(分)', '娱乐时长(分)',
        '上网活动', '心情', '自评(分)', '打卡时间',
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # ── Data rows ──
    for i, record in enumerate(checkins, start=1):
        profile = record.user.profile
        row_data = [
            i,
            profile.phone,
            profile.real_name or profile.nickname,
            profile.school or '-',
            profile.grade or '-',
            record.day,
            record.date.strftime('%Y-%m-%d') if record.date else '-',
            record.study_duration or 0,
            record.entertainment_duration or 0,
            record.online_activities or '-',
            record.mood or '-',
            record.self_rating or '-',
            record.created_at.strftime('%Y-%m-%d %H:%M') if record.created_at else '-',
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i + 1, column=col, value=val)
            cell.alignment = cell_align
            cell.border = thin_border

    # ── Column widths ──
    col_widths = [6, 16, 10, 12, 8, 8, 14, 14, 14, 28, 10, 10, 20]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

    # ── Freeze header ──
    ws.freeze_panes = 'A2'

    # ── Response ──
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    school_label = managed_school if managed_school else '全部'
    filename = f'打卡数据_{school_label}_{date.today().strftime("%Y%m%d")}.xlsx'
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response['Content-Length'] = len(response.content)
    return response
