from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.db.models import Count
from datetime import date
import csv
from openpyxl import load_workbook
from ..models import UserProfile, DailyTopic, CheckInRecord, Message, Reply, Notification, Announcement
from django.contrib.auth.models import User
from .decorators import admin_required, _filter_profiles_by_school, _is_super_admin, _get_managed_school, _load_config, _save_config, DEFAULT_CONFIG


# ═══════════════════════════════════════════════════════════════════════
# 2. User Management
# ═══════════════════════════════════════════════════════════════════════

@admin_required
def admin_users_view(request):
    # ── CSV Export ──
    if request.GET.get('export') == 'csv':
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = 'attachment; filename="users_export.csv"'
        writer = csv.writer(response)
        writer.writerow(['手机号', '昵称', '学校', '年级', '总打卡天数', '完成进度(%)', '最后打卡日期'])

        users_data = _build_users_data(request)
        for u in users_data:
            p = u['profile']
            total = u['total_checkins']
            progress = round(total / 21 * 100, 1)
            writer.writerow([p.phone, p.nickname, p.school, p.grade, total, progress, u['last_checkin_date']])
        return response

    users_data = _build_users_data(request)

    total_users = len(users_data)
    completed_21 = sum(1 for u in users_data if u['total_checkins'] >= 21)
    active_streak = sum(1 for u in users_data if u['streak_days'] >= 7)
    today = date.today().strftime('%Y-%m-%d')
    checked_today = sum(1 for u in users_data if u['last_checkin_date'] == today)

    context = {
        'current_page': 'users',
        'users_data': users_data,
        'total_users': total_users,
        'completed_21': completed_21,
        'active_streak': active_streak,
        'checked_today': checked_today,
        'is_super_admin': _is_super_admin(request.user),
    }
    return render(request, 'admin/users.html', context)


# ═══════════════════════════════════════════════════════════════════════
# 2.5 Excel Import Users
# ═══════════════════════════════════════════════════════════════════════

@admin_required
def admin_import_users(request):
    """Excel 批量导入用户"""
    if request.method != 'POST':
        return redirect('admin_users')

    excel_file = request.FILES.get('file')
    if not excel_file:
        context = {
            'current_page': 'users',
            'users_data': _build_users_data(request),
            'import_error': '请选择一个 Excel 文件',
        }
        return render(request, 'admin/users.html', context)

    # 检查文件扩展名
    filename = excel_file.name.lower()
    if not (filename.endswith('.xlsx') or filename.endswith('.xls')):
        context = {
            'current_page': 'users',
            'users_data': _build_users_data(request),
            'import_error': '仅支持 .xlsx / .xls 格式的 Excel 文件',
        }
        return render(request, 'admin/users.html', context)

    try:
        wb = load_workbook(excel_file, read_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(min_row=2, values_only=True))  # 跳过表头
        imported = 0
        skipped = 0
        errors = []

        for i, row in enumerate(rows, start=2):
            # 读取列：学校, 真实姓名, 电话号
            school = str(row[0]).strip() if row[0] else ''
            real_name = str(row[1]).strip() if row[1] else ''
            phone = str(row[2]).strip() if row[2] else ''

            if not phone or not real_name or not school:
                errors.append(f'第{i}行：学校、姓名、电话号不能为空')
                continue

            # 跳过已存在的手机号
            if UserProfile.objects.filter(phone=phone).exists():
                skipped += 1
                continue

            try:
                user = User.objects.create_user(
                    username=phone,
                    password='123456',
                )
                UserProfile.objects.create(
                    user=user,
                    phone=phone,
                    nickname=real_name,
                    real_name=real_name,
                    school=school,
                    role='user',
                )
                imported += 1
            except Exception as e:
                errors.append(f'第{i}行({real_name})：创建失败 - {str(e)}')

        wb.close()

        msg = f'导入完成：成功 {imported} 人，跳过 {skipped} 人（已存在）'
        if errors:
            msg += f'，{len(errors)} 条错误'

        context = {
            'current_page': 'users',
            'users_data': _build_users_data(request),
            'import_success': msg,
            'import_errors': errors[:20],  # 最多显示 20 条错误
        }
        return render(request, 'admin/users.html', context)

    except Exception as e:
        context = {
            'current_page': 'users',
            'users_data': _build_users_data(request),
            'import_error': f'解析 Excel 文件失败：{str(e)}',
        }
        return render(request, 'admin/users.html', context)


def _build_users_data(request):
    """构建用户数据列表，供 users 页面和导入页面复用
    优化：批量查询所有用户的打卡记录，避免 N+1 查询问题。
    """
    profiles = list(_filter_profiles_by_school(request))
    if not profiles:
        return []

    user_ids = [p.user_id for p in profiles]
    profile_map = {p.user_id: p for p in profiles}

    # 一次性批量获取所有用户的打卡记录
    all_checkins = (
        CheckInRecord.objects
        .filter(user_id__in=user_ids)
        .values('user_id', 'day', 'created_at')
        .order_by('user_id')
    )

    # 按用户分组
    user_checkins = {}
    for c in all_checkins:
        uid = c['user_id']
        if uid not in user_checkins:
            user_checkins[uid] = {'all_days': [], 'last_created_at': None}
        user_checkins[uid]['all_days'].append(c['day'])
        created_at = c['created_at']
        if user_checkins[uid]['last_created_at'] is None or created_at > user_checkins[uid]['last_created_at']:
            user_checkins[uid]['last_created_at'] = created_at

    users_data = []
    seen_uids = set()

    for uid, data in user_checkins.items():
        seen_uids.add(uid)
        total = len(data['all_days'])
        distinct_days = sorted(set(data['all_days']))

        # 计算连续打卡天数（从最近一天往回数）
        streak = 1 if distinct_days else 0
        for j in range(len(distinct_days) - 1, 0, -1):
            if distinct_days[j] == distinct_days[j - 1] + 1:
                streak += 1
            else:
                break

        last_date = data['last_created_at'].strftime('%Y-%m-%d') if data['last_created_at'] else '-'
        users_data.append({
            'profile': profile_map[uid],
            'total_checkins': total,
            'streak_days': streak,
            'last_checkin_date': last_date,
        })

    # 没有打卡记录的用户
    for uid in user_ids:
        if uid not in seen_uids:
            users_data.append({
                'profile': profile_map[uid],
                'total_checkins': 0,
                'streak_days': 0,
                'last_checkin_date': '-',
            })

    return users_data


# ═══════════════════════════════════════════════════════════════════════
# 3. Daily Topic Management
# ═══════════════════════════════════════════════════════════════════════

@admin_required
def admin_topics_view(request):
    if request.method == 'POST':
        day = int(request.POST.get('day', 0))
        if 1 <= day <= 21:
            topic = DailyTopic.objects.filter(day=day).first()
            if topic:
                topic.title = request.POST.get('title', topic.title)
                topic.content = request.POST.get('content', topic.content)
                topic.question = request.POST.get('question', topic.question)
                topic.icon = request.POST.get('icon', topic.icon)
                topic.save()
        return redirect('admin_topics')

    # ── Drill-down by day ──
    day_filter = request.GET.get('day')
    day_records = None
    if day_filter:
        try:
            day_int = int(day_filter)
            allowed_users = _filter_profiles_by_school(request).values_list('user_id', flat=True)
            day_records = CheckInRecord.objects.filter(day=day_int, user_id__in=allowed_users).select_related('user__profile').order_by('-created_at')
        except ValueError:
            pass

    topics = DailyTopic.objects.all().order_by('day')

    # 批量查询每个 topic 的回答数，避免 N+1
    answer_counts = dict(
        CheckInRecord.objects
        .exclude(user__profile__role__in=['admin', 'super_admin'])
        .values('day')
        .annotate(cnt=Count('id'))
        .values_list('day', 'cnt')
    )

    topics_data = []
    for t in topics:
        topics_data.append({
            'topic': t,
            'answer_count': answer_counts.get(t.day, 0),
        })

    context = {
        'current_page': 'topics',
        'topics_data': topics_data,
        'day_records': day_records,
        'filter_day': day_filter,
    }
    return render(request, 'admin/topics.html', context)


# ═══════════════════════════════════════════════════════════════════════
# 4. Message Moderation
# ═══════════════════════════════════════════════════════════════════════

@admin_required
def admin_messages_view(request):
    if request.method == 'POST':
        action = request.POST.get('action', '')
        msg_id = request.POST.get('id', '')

        if msg_id:
            try:
                msg = Message.objects.get(id=int(msg_id))
            except (Message.DoesNotExist, ValueError):
                msg = None

            if msg:
                if action == 'delete':
                    msg.delete()
                elif action == 'hide':
                    if not msg.content.startswith('[已隐藏]'):
                        msg.content = '[已隐藏] ' + msg.content
                        msg.save()

        return redirect('admin_messages')

    # ── GET: list all messages ──
    allowed_users = _filter_profiles_by_school(request).values_list('user_id', flat=True)
    messages = Message.objects.filter(user_id__in=allowed_users).select_related('user__profile').prefetch_related('likes', 'replies').order_by('-created_at')

    msg_list = []
    for m in messages:
        try:
            nickname = m.user.profile.nickname
            phone = m.user.profile.phone
        except UserProfile.DoesNotExist:
            nickname = m.user.username
            phone = '-'
        msg_list.append({
            'message': m,
            'nickname': nickname,
            'phone': phone,
        })

    context = {
        'current_page': 'messages',
        'messages': msg_list,
    }
    return render(request, 'admin/messages.html', context)


# ═══════════════════════════════════════════════════════════════════════
# 5. Rule Configuration
# ═══════════════════════════════════════════════════════════════════════

@admin_required
def admin_config_view(request):
    if request.method == 'POST':
        config = {
            'activity_name': request.POST.get('activity_name', DEFAULT_CONFIG['activity_name']),
            'slogan': request.POST.get('slogan', DEFAULT_CONFIG['slogan']),
            'start_date': request.POST.get('start_date', ''),
            'end_date': request.POST.get('end_date', ''),
            'allow_makeup': request.POST.get('allow_makeup') == 'true',
            'makeup_limit': int(request.POST.get('makeup_limit', DEFAULT_CONFIG['makeup_limit'])),
        }
        _save_config(config)
        return redirect('admin_config')

    context = {
        'current_page': 'config',
        'config': _load_config(),
    }
    return render(request, 'admin/config.html', context)


# ═══════════════════════════════════════════════════════════════════════
# 6. Announcement Publisher
# ═══════════════════════════════════════════════════════════════════════

@admin_required
def admin_announce_view(request):
    """公告管理 - 发布/编辑/删除公告，可见范围与管理员权限一致"""
    from day21.views import create_notification

    # ── DELETE ──
    if request.GET.get('delete'):
        try:
            a = Announcement.objects.get(id=request.GET['delete'])
            a.delete()
        except Announcement.DoesNotExist:
            pass
        return redirect('admin_announce')

    # ── EDIT (load existing) ──
    edit_announcement = None
    if request.GET.get('edit'):
        try:
            edit_announcement = Announcement.objects.get(id=request.GET['edit'])
        except (Announcement.DoesNotExist, ValueError):
            pass

    # ── POST: Create or Update ──
    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        content = request.POST.get('content', '').strip()
        n_type = request.POST.get('n_type', 'announcement')
        edit_id = request.POST.get('edit_id', '')

        if title and content:
            allowed_users_qs = _filter_profiles_by_school(request)
            allowed_users = list(allowed_users_qs.values_list('user_id', flat=True))
            users = User.objects.filter(id__in=allowed_users)

            # 确定目标学校
            managed_school = _get_managed_school(request)
            target_school = managed_school if managed_school and managed_school != '__none__' else ''

            if edit_id:
                # 更新已有公告
                try:
                    a = Announcement.objects.get(id=int(edit_id))
                    a.title = title
                    a.content = content
                    a.n_type = n_type
                    a.target_school = target_school
                    a.recipient_count = len(allowed_users)
                    a.save()
                except Announcement.DoesNotExist:
                    a = Announcement.objects.create(
                        title=title, content=content, n_type=n_type,
                        target_school=target_school,
                        publisher=request.user,
                        recipient_count=len(allowed_users),
                    )
            else:
                a = Announcement.objects.create(
                    title=title, content=content, n_type=n_type,
                    target_school=target_school,
                    publisher=request.user,
                    recipient_count=len(allowed_users),
                )

            for user in users:
                create_notification(user, title, content, n_type)

        return redirect('admin_announce')

    # ── GET: List announcements ──
    managed_school = _get_managed_school(request)
    if managed_school is None:
        announcements = Announcement.objects.all().order_by('-created_at')
        scope_desc = '所有用户'
    elif managed_school == '__none__':
        announcements = Announcement.objects.none()
        scope_desc = '无权限'
    else:
        announcements = Announcement.objects.filter(target_school=managed_school).order_by('-created_at')
        scope_desc = f'学校: {managed_school}'

    context = {
        'current_page': 'announce',
        'announcements': announcements,
        'edit_announcement': edit_announcement,
        'scope_desc': scope_desc,
    }
    return render(request, 'admin/announce.html', context)
